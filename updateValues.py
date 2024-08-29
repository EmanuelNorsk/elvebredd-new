from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import pandas
import json
import os


def getColumnNames(sheet):
    max_column = 1
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                max_column = max(max_column, cell.column)
    column_names = [chr(ord('A') + i) for i in range(max_column)]
    return column_names

def getRowNames(sheet):
    max_row = 1
    for col in sheet.iter_cols():
        for cell in col:
            if cell.value is not None:
                max_row = max(max_row, cell.row)
    row_names = [str(i) for i in range(1, max_row + 1)]
    return row_names

finalDictionary = {}
for fileName in os.listdir("values"):
    print("Loading file: " + str(fileName))
    dictionary = pandas.read_excel("values/" + fileName).to_dict(orient="index")
    wb = load_workbook("values/" + fileName)

    for sheetName in wb.sheetnames:
        print("Loading sheet: " + str(sheetName))
        sheet = wb[sheetName]
        image_loader = SheetImageLoader(sheet)
        print("Extracting data...")
        for columnName in getColumnNames(sheet):
            for rowName in getRowNames(sheet):
                if image_loader.image_in(columnName + rowName):
                    image = image_loader.get(columnName + rowName)
                    image.save("static/images/pets/" + (sheet.cell(int(rowName), 1).value).replace("_", "'") + ".png", "PNG")
        print("Data Extracted from " + str(sheetName))

    print("Adding Image Directories...")
    print("Removing junk data...")
    print("Fixing json convertion artifacts...")

    for key in dictionary.keys():
        dictionary[key]["image"] = "/static/images/pets/" + str(dictionary[key]["Name"]) + ".png"

        delList = []
        for innerKey in dictionary[key].keys():
            if "Unnamed" in innerKey:
                delList.append(innerKey)
            if pandas.isna(dictionary[key][innerKey]):
                dictionary[key][innerKey] = None
        delList.reverse()
        for item in delList:
            dictionary[key].pop(item)

    print("Fixing key names...")

    newDictionary = {}
    for key, value in dictionary.items():
        preDict = {}
        for key2, value2 in value.items():
            preDict[key2.lower()] = value2
        value = preDict
        if value["name"] != None:
            value["id"] = str(int(key) + len(finalDictionary))
            newDictionary[str(int(key) + len(finalDictionary))] = value

    print("Adding data to a dictionary...")

    finalDictionary.update(newDictionary)

    print("File complete: " + str(fileName))

print("All files completed.")
print("Giving each pet a score...")

listingCategories = {}

for key, value in finalDictionary.items():
    petValue = value.get("rvalue", 0)
    if petValue == 0:
        petValue = value.get("value", 0)

    if "type" in value and "rarity" in value:
        if value.get("type", "unknown") + " " + value.get("rarity", "unknown") not in listingCategories:
            listingCategories[value["type"] + " " + value["rarity"]] = [(petValue, key)]
        else:
            listingCategories[value["type"] + " " + value["rarity"]].append((petValue, key))

for key, value in listingCategories.items():
    value.sort(reverse=True)

    for i, tuple in enumerate(value):
        finalDictionary[tuple[1]]["score"] = str(i+1).zfill(3)

print("Done.")
print("Dumping Data...")

with open("data/Pets.json", "w") as file:
    json.dump(finalDictionary, file, indent=4, ensure_ascii=True)

print("Successfully dumped data to: /data/Pets.json")